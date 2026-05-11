"""Generate the Monthly Global Races Report Excel file.

Output matches the 4-sheet structure of the sample workbook:
    1. Race Overview
    2. Elite Results
    3. Sponsorship & Partnerships
    4. Events' Highlights
"""
from __future__ import annotations

import calendar
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app import data
from app.scrapers.base import RaceFacts, time_to_str
from app.scrapers.registry import scrape_race

ROOT = Path(__file__).resolve().parents[1]
REPORTS_DIR = ROOT / "data" / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

# --- styling tokens ---------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F3864")  # deep navy
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
ZEBRA_FILL = PatternFill("solid", fgColor="F2F6FB")
ACCENT_FILL = PatternFill("solid", fgColor="C00000")
SECTION_FILL = PatternFill("solid", fgColor="305496")

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10)
BODY_BOLD = Font(name="Calibri", size=10, bold=True)

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _style_header(ws, row: int, cols: range, fill=HEADER_FILL, font=HEADER_FONT) -> None:
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = BORDER


def _format_location(venue: str, country: str) -> str:
    venue = (venue or "").strip()
    country = (country or "").strip()
    if not country:
        return venue
    if "(" in venue:
        return venue
    return f"{venue} ({country})"


def _format_category(cat: str) -> str:
    mapping = {"Marathon": "M", "Half Marathon": "HM", "10K": "10K"}
    return mapping.get(cat, cat)


# ---------------------------------------------------------------------------
# Sheet 1 — Race Overview
# ---------------------------------------------------------------------------
def _write_race_overview(wb: Workbook, races, facts_by_id: dict[str, RaceFacts]) -> None:
    ws = wb.create_sheet("Race Overview")
    headers = [
        "Ser. No.", "Date", "Event", "Location", "Inception Year", "Edition",
        "WA Label", "Race Category", "Finishers (Feature Race only)", "", "", "",
        "Spectators", "Volunteers", "Prize Money (USD)",
    ]
    sub = ["", "", "", "", "", "", "", "", "Numbers", "Men%", "Women%", "Non Binary%", "", "", ""]

    # Row 1 left blank to mirror the original sample
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    for c, h in enumerate(sub, start=1):
        if h:
            ws.cell(row=3, column=c, value=h)

    ws.merge_cells(start_row=2, start_column=9, end_row=2, end_column=12)  # Finishers group
    for col in (1, 2, 3, 4, 5, 6, 7, 8, 13, 14, 15):
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)

    _style_header(ws, 2, range(1, 16))
    _style_header(ws, 3, range(9, 13), fill=SUBHEADER_FILL, font=SUBHEADER_FONT)

    row = 4
    for i, race in enumerate(races, start=1):
        f = facts_by_id.get(race.id) or RaceFacts(race_id=race.id)
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=race.date).number_format = "dd-mmm-yyyy"
        ws.cell(row=row, column=3, value=race.name)
        ws.cell(row=row, column=4, value=_format_location(race.venue, race.country))
        ws.cell(row=row, column=5, value=f.inception_year)
        ws.cell(row=row, column=6, value=f.edition)
        ws.cell(row=row, column=7, value=race.label)
        ws.cell(row=row, column=8, value=_format_category(race.category))
        ws.cell(row=row, column=9, value=f.finishers_total)
        ws.cell(row=row, column=10, value=f.finishers_men_pct)
        ws.cell(row=row, column=11, value=f.finishers_women_pct)
        ws.cell(row=row, column=12, value=f.finishers_nonbinary_pct)
        ws.cell(row=row, column=13, value=f.spectators)
        ws.cell(row=row, column=14, value=f.volunteers)
        ws.cell(row=row, column=15, value=f.prize_money_usd)

        for c in range(1, 16):
            cell = ws.cell(row=row, column=c)
            cell.font = BODY_FONT
            cell.alignment = CENTER if c in (1, 2, 5, 6, 7, 8, 9, 10, 11, 12) else LEFT
            cell.border = BORDER
            if i % 2 == 0:
                cell.fill = ZEBRA_FILL

        for c in (9, 13, 14, 15):
            ws.cell(row=row, column=c).number_format = "#,##0"
        for c in (10, 11, 12):
            ws.cell(row=row, column=c).number_format = "0.0"
        row += 1

    widths = [9, 12, 36, 24, 12, 9, 11, 14, 11, 9, 9, 11, 12, 11, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Sheet 2 — Elite Results
# ---------------------------------------------------------------------------
WORLD_RECORDS = [
    ("Marathon", "Men", "2:00:35", "Kelvin Kiptum (KEN)", "Chicago 2023"),
    ("Marathon", "Womens (Mixed)", "2:09:56", "Ruth Chepngetich (KEN)", "Chicago 2024"),
    ("Marathon", "Women", "2:15:50", "Tigst Assefa (ETH)", "London 2025"),
    ("Half Marathon", "Men", "0:57:20", "Jacob Kiplimo (UGA)", "Lisbon 2026"),
    ("Half Marathon", "Womens (Mixed)", "1:02:52", "Letsenbet Gidey (ETH)", "Valencia 2021"),
    ("Half Marathon", "Women", "1:05:16", "Peres Jepchirchir (KEN)", "Gydnia 2020"),
]

LEGEND = [("WR", "World Record"), ("CR", "Course Record"), ("NR", "National Record")]


def _write_elite_results(wb: Workbook, races, facts_by_id: dict[str, RaceFacts]) -> None:
    ws = wb.create_sheet("Elite Results")

    # Top: World Records block (cols B..F) + legend (cols I..J)
    ws.cell(row=2, column=2, value="World Records").font = HEADER_FONT
    ws.cell(row=2, column=2).fill = HEADER_FILL
    ws.cell(row=2, column=2).alignment = LEFT
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=6)

    cur_distance = None
    r = 3
    for distance, gender, time_str, holder, where in WORLD_RECORDS:
        if distance != cur_distance:
            ws.cell(row=r, column=2, value=distance).font = BODY_BOLD
            cur_distance = distance
        ws.cell(row=r, column=3, value=gender)
        ws.cell(row=r, column=4, value=time_str)
        ws.cell(row=r, column=5, value=holder)
        ws.cell(row=r, column=6, value=where)
        for c in range(2, 7):
            cell = ws.cell(row=r, column=c)
            if c != 2:  # leave the bold distance label alone
                cell.font = BODY_FONT
            cell.alignment = LEFT
            cell.border = BORDER
        r += 1

    for i, (abbr, meaning) in enumerate(LEGEND):
        ws.cell(row=3 + i, column=9, value=abbr).font = BODY_BOLD
        ws.cell(row=3 + i, column=10, value=meaning).font = BODY_FONT

    # Podium section
    section_row = 10
    headers = [
        "", "Date", "Event", "Rank",
        "Men's Podium", "", "", "",
        "",
        "Women's Podium", "", "", "",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=section_row, column=c, value=h)
    sub = ["", "", "", "", "Name", "Nationality", "Timing", "Remark", "", "Name", "Nationality", "Timing", "Remark"]
    for c, h in enumerate(sub, start=1):
        ws.cell(row=section_row + 1, column=c, value=h)

    ws.merge_cells(start_row=section_row, start_column=5, end_row=section_row, end_column=8)
    ws.merge_cells(start_row=section_row, start_column=10, end_row=section_row, end_column=13)
    for col in (2, 3, 4):
        ws.merge_cells(start_row=section_row, start_column=col, end_row=section_row + 1, end_column=col)

    _style_header(ws, section_row, range(2, 14))
    _style_header(ws, section_row + 1, range(5, 14), fill=SUBHEADER_FILL, font=SUBHEADER_FONT)

    row = section_row + 2
    for race in races:
        f = facts_by_id.get(race.id) or RaceFacts(race_id=race.id)
        # short event label like "Tokyo Marathon" -> use full name
        for rank in (1, 2, 3):
            men = next((p for p in f.mens_podium if p.rank == rank), None)
            women = next((p for p in f.womens_podium if p.rank == rank), None)
            if rank == 1:
                ws.cell(row=row, column=2, value=race.date).number_format = "dd-mmm-yyyy"
                ws.cell(row=row, column=3, value=race.name)
            ws.cell(row=row, column=4, value=rank)
            if men:
                ws.cell(row=row, column=5, value=men.name)
                ws.cell(row=row, column=6, value=men.nationality)
                ws.cell(row=row, column=7, value=time_to_str(men.timing))
                ws.cell(row=row, column=8, value=men.remark)
            if women:
                ws.cell(row=row, column=10, value=women.name)
                ws.cell(row=row, column=11, value=women.nationality)
                ws.cell(row=row, column=12, value=time_to_str(women.timing))
                ws.cell(row=row, column=13, value=women.remark)
            for c in range(2, 14):
                cell = ws.cell(row=row, column=c)
                cell.font = BODY_FONT
                cell.alignment = LEFT_TOP if c in (3, 5, 8, 10, 13) else CENTER
                cell.border = BORDER
            row += 1
        row += 1  # spacer

    widths = [3, 12, 28, 7, 22, 12, 12, 26, 3, 22, 12, 12, 26]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{section_row + 2}"


# ---------------------------------------------------------------------------
# Sheet 3 — Sponsorship & Partnerships
# ---------------------------------------------------------------------------
def _write_sponsorship(wb: Workbook, races, facts_by_id: dict[str, RaceFacts]) -> None:
    ws = wb.create_sheet("Sponsorship & Partnerships")
    headers = ["", "Date", "Event", "Location", "Organizers / Promoters", "Title Sponsor", "Other Sponsors/ Partners"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    _style_header(ws, 2, range(2, 8))

    row = 3
    for i, race in enumerate(races, start=1):
        f = facts_by_id.get(race.id) or RaceFacts(race_id=race.id)
        ws.cell(row=row, column=2, value=race.date).number_format = "dd-mmm-yyyy"
        ws.cell(row=row, column=3, value=race.name)
        ws.cell(row=row, column=4, value=_format_location(race.venue, race.country))
        ws.cell(row=row, column=5, value=f.organizers)
        ws.cell(row=row, column=6, value=f.title_sponsor)
        ws.cell(row=row, column=7, value=f.other_sponsors)
        for c in range(2, 8):
            cell = ws.cell(row=row, column=c)
            cell.font = BODY_FONT
            cell.alignment = LEFT_TOP
            cell.border = BORDER
            if i % 2 == 0:
                cell.fill = ZEBRA_FILL
        row += 1

    widths = [3, 12, 32, 22, 30, 22, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Sheet 4 — Events' Highlights
# ---------------------------------------------------------------------------
def _write_highlights(wb: Workbook, races, facts_by_id: dict[str, RaceFacts]) -> None:
    ws = wb.create_sheet("Events' Highlights")
    headers = ["", "Date", "Event", "Location", "Highlight 1", "Highlight 2", "Highlight 3", "Highlight 4", "Highlight 5"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    _style_header(ws, 2, range(2, 10))

    row = 3
    for i, race in enumerate(races, start=1):
        f = facts_by_id.get(race.id) or RaceFacts(race_id=race.id)
        ws.cell(row=row, column=2, value=race.date).number_format = "dd-mmm-yyyy"
        ws.cell(row=row, column=3, value=race.name)
        ws.cell(row=row, column=4, value=_format_location(race.venue, race.country))
        for j in range(5):
            col = 5 + j
            if j < len(f.highlights):
                title, url = f.highlights[j]
                cell = ws.cell(row=row, column=col, value=f"{title} ({url})" if url else title)
            else:
                ws.cell(row=row, column=col, value="")
        for c in range(2, 10):
            cell = ws.cell(row=row, column=c)
            cell.font = BODY_FONT
            cell.alignment = LEFT_TOP
            cell.border = BORDER
            if i % 2 == 0:
                cell.fill = ZEBRA_FILL
        row += 1

    widths = [3, 12, 32, 22, 40, 40, 40, 40, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Public entrypoint
# ---------------------------------------------------------------------------
def generate_report(year: int, month: int) -> Path:
    races = data.races_for_month(year, month)
    if not races:
        raise ValueError(f"No shortlisted races for {year}-{month:02d}")

    facts_by_id: dict[str, RaceFacts] = {}
    for r in races:
        facts_by_id[r.id] = scrape_race(r.id, r.official_url)

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet
    _write_race_overview(wb, races, facts_by_id)
    _write_elite_results(wb, races, facts_by_id)
    _write_sponsorship(wb, races, facts_by_id)
    _write_highlights(wb, races, facts_by_id)

    month_name = calendar.month_name[month]
    today = datetime.now().strftime("%d.%m.%Y")
    fname = f"Monthly Global Races Report_{month_name} {year}_{today}.xlsx"
    out = REPORTS_DIR / fname
    wb.save(out)
    return out


if __name__ == "__main__":
    p = generate_report(2026, 3)
    print(f"Wrote {p}")
