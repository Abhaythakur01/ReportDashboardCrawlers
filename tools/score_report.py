"""Score a generated monthly report on two axes:

1. **Format compliance** vs. the original sample template
   - Sheet names, header strings, column widths, merged-cell groups,
     row offsets where data starts.
2. **Data depth** per race
   - For each race in the report, what fraction of the expected fields
     are populated? Combined into per-race + overall coverage.

Usage:
    python -m tools.score_report data/reports/...April...xlsx \
                                "Monthly Global Races Report_March 2026_14.04.2026.xlsx"
"""
from __future__ import annotations

import sys
from datetime import datetime, time
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Format-compliance check — does the generated file match the template?
# ---------------------------------------------------------------------------

EXPECTED_SHEETS = ["Race Overview", "Elite Results", "Sponsorship & Partnerships", "Events' Highlights"]

OVERVIEW_HEADERS = [
    "Ser. No.", "Date", "Event", "Location", "Inception Year", "Edition",
    "WA Label", "Race Category", "Finishers (Feature Race only)", None,
    None, None, "Spectators", "Volunteers", "Prize Money (USD)",
]
OVERVIEW_SUBHEADERS_GROUP = ["Numbers", "Men%", "Women%", "Non Binary%"]

SPONSORSHIP_HEADERS = [
    None, "Date", "Event", "Location", "Organizers / Promoters",
    "Title Sponsor", "Other Sponsors/ Partners",
]

HIGHLIGHTS_HEADERS = [
    None, "Date", "Event", "Location",
    "Highlight 1", "Highlight 2", "Highlight 3", "Highlight 4", "Highlight 5",
]


def _norm_header(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def check_format(ours_path: Path, orig_path: Path) -> dict:
    ours = openpyxl.load_workbook(ours_path, data_only=True)
    orig = openpyxl.load_workbook(orig_path, data_only=True)

    checks: list[tuple[str, bool, str]] = []

    # 1. Sheet names
    sheets_match = ours.sheetnames == orig.sheetnames
    checks.append(("Sheet names match exactly", sheets_match,
                   f"ours={ours.sheetnames!r} orig={orig.sheetnames!r}" if not sheets_match else ""))

    # 2. Race Overview headers (row 2)
    o_ws = ours["Race Overview"]
    g_ws = orig["Race Overview"]
    for c, exp in enumerate(OVERVIEW_HEADERS, start=1):
        if exp is None:
            continue
        ours_h = _norm_header(o_ws.cell(row=2, column=c).value)
        orig_h = _norm_header(g_ws.cell(row=2, column=c).value)
        ok = ours_h == orig_h
        checks.append((f"Race Overview header col{c}: {exp!r}", ok,
                       f"ours={ours_h!r} orig={orig_h!r}" if not ok else ""))

    # 3. Race Overview sub-headers under "Finishers" (row 3, cols 9-12)
    for i, exp in enumerate(OVERVIEW_SUBHEADERS_GROUP):
        c = 9 + i
        ours_h = _norm_header(o_ws.cell(row=3, column=c).value)
        orig_h = _norm_header(g_ws.cell(row=3, column=c).value)
        ok = ours_h == orig_h
        checks.append((f"Race Overview sub-header col{c}: {exp!r}", ok,
                       f"ours={ours_h!r} orig={orig_h!r}" if not ok else ""))

    # 4. Race Overview merged cells — Finishers should span cols 9-12 on row 2
    o_merged = {str(r) for r in o_ws.merged_cells.ranges}
    g_merged = {str(r) for r in g_ws.merged_cells.ranges}
    finishers_merge = "I2:L2"  # cols 9-12 on row 2
    ok = finishers_merge in o_merged and finishers_merge in g_merged
    checks.append((f"Race Overview Finishers merged ({finishers_merge})", ok, ""))

    # 5. Sponsorship headers (row 2)
    o_ws = ours["Sponsorship & Partnerships"]
    g_ws = orig["Sponsorship & Partnerships"]
    for c, exp in enumerate(SPONSORSHIP_HEADERS, start=1):
        if exp is None:
            continue
        ours_h = _norm_header(o_ws.cell(row=2, column=c).value)
        orig_h = _norm_header(g_ws.cell(row=2, column=c).value)
        ok = ours_h == orig_h
        checks.append((f"Sponsorship header col{c}: {exp!r}", ok,
                       f"ours={ours_h!r} orig={orig_h!r}" if not ok else ""))

    # 6. Highlights headers (row 2)
    o_ws = ours["Events' Highlights"]
    g_ws = orig["Events' Highlights"]
    for c, exp in enumerate(HIGHLIGHTS_HEADERS, start=1):
        if exp is None:
            continue
        ours_h = _norm_header(o_ws.cell(row=2, column=c).value)
        orig_h = _norm_header(g_ws.cell(row=2, column=c).value)
        ok = ours_h == orig_h
        checks.append((f"Highlights header col{c}: {exp!r}", ok,
                       f"ours={ours_h!r} orig={orig_h!r}" if not ok else ""))

    passed = sum(1 for _, ok, _ in checks if ok)
    total = len(checks)
    return {
        "checks": checks,
        "passed": passed,
        "total": total,
        "pct": passed / total * 100 if total else 0,
    }


# ---------------------------------------------------------------------------
# Data-depth scoring — per race, what fraction of expected fields are filled?
# ---------------------------------------------------------------------------

OVERVIEW_DATA_FIELDS = [
    ("inception_year", 5),
    ("edition", 6),
    ("finishers_total", 9),
    ("finishers_men_pct", 10),
    ("finishers_women_pct", 11),
    ("spectators", 13),
    ("volunteers", 14),
    ("prize_money", 15),
]

SPONSORSHIP_DATA_FIELDS = [
    ("organizers", 5),
    ("title_sponsor", 6),
    ("other_sponsors", 7),
]


def _is_filled(v) -> bool:
    if v is None:
        return False
    if isinstance(v, str) and not v.strip():
        return False
    return True


def score_data_depth(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    races: dict[str, dict] = {}

    # Race Overview
    ws = wb["Race Overview"]
    for r in range(4, ws.max_row + 1):
        ev = ws.cell(row=r, column=3).value
        if not ev:
            continue
        race = races.setdefault(str(ev).strip(), {"overview": {}, "sponsorship": {}, "highlights": [], "podium": []})
        for fld, col in OVERVIEW_DATA_FIELDS:
            race["overview"][fld] = ws.cell(row=r, column=col).value

    # Sponsorship
    ws = wb["Sponsorship & Partnerships"]
    for r in range(3, ws.max_row + 1):
        ev = ws.cell(row=r, column=3).value
        if not ev:
            continue
        race = races.setdefault(str(ev).strip(), {"overview": {}, "sponsorship": {}, "highlights": [], "podium": []})
        for fld, col in SPONSORSHIP_DATA_FIELDS:
            race["sponsorship"][fld] = ws.cell(row=r, column=col).value

    # Highlights
    ws = wb["Events' Highlights"]
    for r in range(3, ws.max_row + 1):
        ev = ws.cell(row=r, column=3).value
        if not ev:
            continue
        race = races.setdefault(str(ev).strip(), {"overview": {}, "sponsorship": {}, "highlights": [], "podium": []})
        for c in range(5, 10):
            race["highlights"].append(ws.cell(row=r, column=c).value)

    # Elite Results — group rank=1 with following rank=2,3 rows
    ws = wb["Elite Results"]
    cur = None
    for r in range(1, ws.max_row + 1):
        rank = ws.cell(row=r, column=4).value
        ev = ws.cell(row=r, column=3).value
        if isinstance(rank, int) and rank == 1 and ev:
            cur = str(ev).strip()
            races.setdefault(cur, {"overview": {}, "sponsorship": {}, "highlights": [], "podium": []})
        if isinstance(rank, int) and rank in (1, 2, 3) and cur is not None:
            men_n = ws.cell(row=r, column=5).value
            men_t = ws.cell(row=r, column=7).value
            wom_n = ws.cell(row=r, column=10).value
            wom_t = ws.cell(row=r, column=12).value
            races[cur]["podium"].append({"rank": rank, "men": (men_n, men_t), "women": (wom_n, wom_t)})

    # Score per race
    per_race = []
    for name, data in races.items():
        ov_total = len(OVERVIEW_DATA_FIELDS)
        ov_filled = sum(1 for v in data["overview"].values() if _is_filled(v))
        sp_total = len(SPONSORSHIP_DATA_FIELDS)
        sp_filled = sum(1 for v in data["sponsorship"].values() if _is_filled(v))
        h_total = 5
        h_filled = sum(1 for v in data["highlights"][:5] if _is_filled(v))
        # Podium: 6 expected (rank 1-3, men + women)
        pod_filled = 0
        for pod in data["podium"]:
            if _is_filled(pod["men"][0]) and _is_filled(pod["men"][1]):
                pod_filled += 1
            if _is_filled(pod["women"][0]) and _is_filled(pod["women"][1]):
                pod_filled += 1
        pod_total = 6

        total = ov_total + sp_total + h_total + pod_total
        filled = ov_filled + sp_filled + h_filled + pod_filled
        per_race.append({
            "race": name,
            "overview": (ov_filled, ov_total),
            "sponsorship": (sp_filled, sp_total),
            "highlights": (h_filled, h_total),
            "podium": (pod_filled, pod_total),
            "total": (filled, total),
            "pct": filled / total * 100 if total else 0,
        })

    return {
        "per_race": per_race,
        "race_count": len(races),
        "overall_filled": sum(r["total"][0] for r in per_race),
        "overall_total": sum(r["total"][1] for r in per_race),
    }


# ---------------------------------------------------------------------------
def main(ours_path: Path, orig_path: Path) -> None:
    print(f"Generated:  {ours_path.name}")
    print(f"Template:   {orig_path.name}")
    print()

    # Format check
    fmt = check_format(ours_path, orig_path)
    print("=" * 78)
    print(f"FORMAT COMPLIANCE   {fmt['passed']}/{fmt['total']} checks pass ({fmt['pct']:.1f}%)")
    print("=" * 78)
    failures = [(name, msg) for name, ok, msg in fmt["checks"] if not ok]
    if failures:
        print("Failed checks:")
        for name, msg in failures:
            print(f"  ✗ {name}")
            if msg:
                print(f"      {msg}")
    else:
        print("  All format checks passed.")
    print()

    # Data depth
    depth = score_data_depth(ours_path)
    print("=" * 78)
    print(f"DATA DEPTH   {depth['overall_filled']}/{depth['overall_total']} fields filled "
          f"({depth['overall_filled'] / depth['overall_total'] * 100 if depth['overall_total'] else 0:.1f}%)"
          f"   across {depth['race_count']} races")
    print("=" * 78)
    print(f"{'RACE':46} {'OV':>5} {'SP':>5} {'HL':>5} {'PD':>5}  {'TOTAL':>9}  {'%':>5}")
    print("-" * 78)
    depth["per_race"].sort(key=lambda r: -r["pct"])
    for rec in depth["per_race"]:
        ov = f"{rec['overview'][0]}/{rec['overview'][1]}"
        sp = f"{rec['sponsorship'][0]}/{rec['sponsorship'][1]}"
        hl = f"{rec['highlights'][0]}/{rec['highlights'][1]}"
        pd = f"{rec['podium'][0]}/{rec['podium'][1]}"
        tot = f"{rec['total'][0]}/{rec['total'][1]}"
        print(f"{rec['race'][:46]:46} {ov:>5} {sp:>5} {hl:>5} {pd:>5}  {tot:>9}  {rec['pct']:>4.0f}%")
    print()
    print("Legend: OV=Race Overview · SP=Sponsorship · HL=Highlights · PD=Podium")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        root = Path(__file__).resolve().parent.parent
        candidates = sorted((root / "data" / "reports").glob("*April*.xlsx"),
                            key=lambda p: p.stat().st_mtime, reverse=True)
        ours = candidates[0]
        orig = root / "Monthly Global Races Report_March 2026_14.04.2026.xlsx"
    else:
        ours = Path(sys.argv[1])
        orig = Path(sys.argv[2])
    main(ours, orig)
