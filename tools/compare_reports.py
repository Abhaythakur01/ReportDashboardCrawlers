"""Compare a generated monthly report against the original sample.

Reads both workbooks, lines up rows by race name, then scores:
  * coverage   — % of fields where we produced any value
  * accuracy   — among populated fields, % that match the original
  * per-race + per-sheet breakdown

Usage:
    python -m tools.compare_reports OURS.xlsx ORIG.xlsx
"""
from __future__ import annotations

import sys
from datetime import datetime, time
from pathlib import Path

import openpyxl


def norm(v):
    """Normalise cell value for comparison."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, time):
        return v.isoformat(timespec="seconds")
    if isinstance(v, str):
        s = v.strip().replace("\xa0", " ").replace("’", "'")
        # Whitespace collapse
        s = " ".join(s.split())
        return s.lower() if s else None
    if isinstance(v, float):
        return round(v, 2)
    return v


def name_norm(s) -> str:
    if not s:
        return ""
    s = str(s).strip().replace("’", "'")
    s = " ".join(s.split())
    return s.lower()


def find_header_row(ws):
    for r in range(1, 6):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str):
                low = v.strip().lower()
                if low in {"ser. no.", "date", "event"}:
                    return r
    return 2


# ---------------------------------------------------------------------------
def load_overview(wb) -> dict[str, dict]:
    ws = wb["Race Overview"]
    rows: dict[str, dict] = {}
    headers = [
        "ser_no", "date", "event", "location", "inception_year", "edition",
        "wa_label", "race_category", "finishers_total", "men_pct",
        "women_pct", "nonbinary_pct", "spectators", "volunteers", "prize_money",
    ]
    # Data starts at row 4 in our format (row 1 blank, 2-3 header)
    for r in range(4, ws.max_row + 1):
        rec = {h: ws.cell(row=r, column=i + 1).value for i, h in enumerate(headers)}
        ev = name_norm(rec.get("event"))
        if ev:
            rows[ev] = rec
    return rows


def load_sponsorship(wb) -> dict[str, dict]:
    ws = wb["Sponsorship & Partnerships"]
    rows: dict[str, dict] = {}
    # Headers at row 2, data from row 3, columns are at offset 1 (col B onward)
    for r in range(3, ws.max_row + 1):
        ev = name_norm(ws.cell(row=r, column=3).value)
        if not ev:
            continue
        rows[ev] = {
            "date": ws.cell(row=r, column=2).value,
            "location": ws.cell(row=r, column=4).value,
            "organizers": ws.cell(row=r, column=5).value,
            "title_sponsor": ws.cell(row=r, column=6).value,
            "other_sponsors": ws.cell(row=r, column=7).value,
        }
    return rows


def load_highlights(wb) -> dict[str, dict]:
    ws = wb["Events' Highlights"]
    rows: dict[str, dict] = {}
    for r in range(3, ws.max_row + 1):
        ev = name_norm(ws.cell(row=r, column=3).value)
        if not ev:
            continue
        rows[ev] = {
            f"highlight_{i}": ws.cell(row=r, column=4 + i).value
            for i in range(1, 6)
        }
    return rows


def load_podium(wb) -> dict[str, dict]:
    """Group Elite Results into per-race {men_1..men_3, women_1..women_3}."""
    ws = wb["Elite Results"]
    rows: dict[str, dict] = {}
    current = None
    for r in range(1, ws.max_row + 1):
        rank = ws.cell(row=r, column=4).value
        ev = ws.cell(row=r, column=3).value
        if isinstance(rank, int) and rank in (1, 2, 3):
            if rank == 1 and ev:
                current = name_norm(ev)
                rows[current] = {}
            if current is None:
                continue
            men_n = ws.cell(row=r, column=5).value
            men_c = ws.cell(row=r, column=6).value
            men_t = ws.cell(row=r, column=7).value
            wom_n = ws.cell(row=r, column=10).value
            wom_c = ws.cell(row=r, column=11).value
            wom_t = ws.cell(row=r, column=12).value
            rows[current][f"men_{rank}"] = (men_n, men_c, men_t)
            rows[current][f"women_{rank}"] = (wom_n, wom_c, wom_t)
    return rows


# ---------------------------------------------------------------------------
_STOPWORDS = {
    # Sponsor / title prefixes that appear inconsistently in race names
    "generali", "zurich", "hoka", "edp", "united", "airlines", "tcs",
    "nn", "schneider", "electric", "baa", "spd", "bank", "of", "china",
    "marato", "marathon", "half", "halbmarathon", "marato", "semi", "de",
    "the", "city", "international", "presented", "by", "asics",
    # Time-keeping / partner prefixes
    "marato",
}

# Specific equivalence pairs (orig_name → our_name). Beats fuzzy matching
# for the cases where naming diverges sharply.
_NAME_ALIASES = {
    "berlin half": "generali berliner halbmarathon",
    "berlin half marathon": "generali berliner halbmarathon",
    "generali berlin half marathon": "generali berliner halbmarathon",
    "barcelona marathon": "zurich marato de barcelona",
    "hoka paris half": "hoka semi de paris",
    "edp lisbon half": "edp lisbon half marathon",
    "meishan renshou half": "meishan renshou half marathon",
    "wan jin shi marathon": "new taipei city wan jin shi marathon",
    "shanghai half": "shanghai half marathon",
    "spd bank shanghai half marathon": "shanghai half marathon",
    "nyc half": "united airlines nyc half",
    "nyc half marathon": "united airlines nyc half",
    "united airlines nyc half marathon": "united airlines nyc half",
    "prague half": "generali prague half marathon",
    "prague half marathon": "generali prague half marathon",
    "yangzhou half": "yangzhou half marathon",
    "bank of china yangzhou half marathon": "yangzhou half marathon",
    "nagoya women's": "nagoya women's marathon",
}


def _key_terms(name: str) -> set[str]:
    name = name.lower().replace("-", " ").replace("'", "").replace("’", "")
    return {t for t in name.split() if t and t not in _STOPWORDS and len(t) > 2}


def fuzzy_match_name(ours_keys, orig_key):
    """Match original race name to one of our keys.

    Tries: exact match → curated alias table → high-overlap key-term match
    after stripping sponsor / title noise.
    """
    keys = list(ours_keys)
    if orig_key in keys:
        return orig_key

    aliased = _NAME_ALIASES.get(orig_key)
    if aliased and aliased in keys:
        return aliased

    orig_terms = _key_terms(orig_key)
    if not orig_terms:
        return None

    best = None
    best_score = 0.0
    for k in keys:
        kt = _key_terms(k)
        if not kt:
            continue
        overlap = len(orig_terms & kt)
        if overlap == 0:
            continue
        # Jaccard against the smaller set of orig terms — favour our_key
        # when it includes all of orig's distinctive terms even if it has
        # extra ones (e.g. "wan jin shi marathon" ⊂ "new taipei city wan
        # jin shi marathon").
        score = overlap / len(orig_terms)
        if score > best_score and score >= 0.5:
            best_score, best = score, k
    return best


def _is_effectively_none(v) -> bool:
    if v is None:
        return True
    if isinstance(v, tuple):
        return all(x is None or (isinstance(x, str) and not x.strip()) for x in v)
    if isinstance(v, str) and not v.strip():
        return True
    return False


def compare_field(ours, orig) -> tuple[bool, str]:
    """Return (matches, status) where status ∈ {'match','mismatch','missing','blank'}."""
    if _is_effectively_none(orig):
        return True, "blank"
    if _is_effectively_none(ours):
        return False, "missing"
    o = norm(ours)
    g = norm(orig)
    if g is None:
        return True, "blank"
    if o is None:
        return False, "missing"
    # Podium triples: (name, country, time)
    if isinstance(ours, tuple) and isinstance(orig, tuple):
        on = name_norm(ours[0])
        gn = name_norm(orig[0])
        ot = norm(ours[2])
        gt = norm(orig[2])
        # Match if either name OR time matches — recap timings/names sometimes
        # diverge slightly between sources but agreement on either is a hit.
        if on and gn and (on == gn or (ot and gt and ot == gt)):
            return True, "match"
        if on and gn and (on in gn or gn in on):
            return True, "partial(name~)"
        return False, "mismatch"

    # For long sponsor strings, treat partial overlap as a partial match
    if isinstance(o, str) and isinstance(g, str) and len(g) > 20:
        if o == g:
            return True, "match"
        # Token-set Jaccard
        ours_tokens = set(o.replace("\n", " ").replace(",", " ").split())
        orig_tokens = set(g.replace("\n", " ").replace(",", " ").split())
        if not orig_tokens:
            return True, "match"
        overlap = len(ours_tokens & orig_tokens) / len(orig_tokens)
        if overlap >= 0.5:
            return True, f"partial({overlap:.0%})"
        return False, "mismatch"
    return (o == g), ("match" if o == g else "mismatch")


# ---------------------------------------------------------------------------
def score(ours_path: Path, orig_path: Path) -> None:
    ours_wb = openpyxl.load_workbook(ours_path, data_only=True)
    orig_wb = openpyxl.load_workbook(orig_path, data_only=True)

    sections = {
        "Race Overview": (load_overview(ours_wb), load_overview(orig_wb),
                          ["inception_year", "edition", "finishers_total",
                           "men_pct", "women_pct", "nonbinary_pct",
                           "spectators", "volunteers", "prize_money"]),
        "Sponsorship": (load_sponsorship(ours_wb), load_sponsorship(orig_wb),
                        ["organizers", "title_sponsor", "other_sponsors"]),
        "Highlights": (load_highlights(ours_wb), load_highlights(orig_wb),
                       [f"highlight_{i}" for i in range(1, 6)]),
        "Elite Results": (load_podium(ours_wb), load_podium(orig_wb),
                          [f"men_{i}" for i in (1, 2, 3)] + [f"women_{i}" for i in (1, 2, 3)]),
    }

    # Aggregate stats
    section_summary: dict[str, dict] = {}
    per_race_summary: dict[str, dict] = {}

    for sect_name, (ours, orig, fields) in sections.items():
        sect = {"populated": 0, "expected": 0, "match": 0, "mismatch": 0, "missing": 0, "blank": 0, "partial": 0}
        for orig_race, orig_row in orig.items():
            our_key = fuzzy_match_name(ours.keys(), orig_race)
            our_row = ours.get(our_key, {}) if our_key else {}
            race_stats = per_race_summary.setdefault(orig_race, {"hit": 0, "miss": 0, "blank": 0, "partial": 0, "details": []})

            for f in fields:
                orig_val = orig_row.get(f)
                our_val = our_row.get(f) if our_row else None
                ok, status = compare_field(our_val, orig_val)
                if status == "blank":
                    sect["blank"] += 1
                    race_stats["blank"] += 1
                    continue
                sect["expected"] += 1
                if status == "missing":
                    sect["missing"] += 1
                    race_stats["miss"] += 1
                else:
                    sect["populated"] += 1
                    if status == "match":
                        sect["match"] += 1
                        race_stats["hit"] += 1
                    elif status.startswith("partial"):
                        sect["partial"] += 1
                        race_stats["partial"] += 1
                    else:
                        sect["mismatch"] += 1
                        race_stats["miss"] += 1

        section_summary[sect_name] = sect

    # Print report
    print(f"\nGenerated:  {ours_path.name}")
    print(f"Reference:  {orig_path.name}")
    print("=" * 78)
    print(f"{'SECTION':28} {'EXPECTED':>9} {'POPULATED':>10} {'MATCH':>7} {'PARTIAL':>8} {'MISMATCH':>9} {'MISSING':>8}")
    print("-" * 78)
    grand_exp = grand_pop = grand_match = grand_partial = grand_miss = grand_missing = 0
    for sect_name, s in section_summary.items():
        print(f"{sect_name:28} {s['expected']:>9} {s['populated']:>10} {s['match']:>7} {s['partial']:>8} {s['mismatch']:>9} {s['missing']:>8}")
        grand_exp += s["expected"]
        grand_pop += s["populated"]
        grand_match += s["match"]
        grand_partial += s["partial"]
        grand_miss += s["mismatch"]
        grand_missing += s["missing"]
    print("-" * 78)
    print(f"{'TOTAL':28} {grand_exp:>9} {grand_pop:>10} {grand_match:>7} {grand_partial:>8} {grand_miss:>9} {grand_missing:>8}")
    print()
    if grand_exp:
        print(f"  Coverage  (populated / expected):                 {grand_pop / grand_exp * 100:5.1f}%")
        print(f"  Accuracy  (match / populated):                    {grand_match / grand_pop * 100 if grand_pop else 0:5.1f}%")
        print(f"  Accuracy+ (match+partial / populated):            {(grand_match + grand_partial) / grand_pop * 100 if grand_pop else 0:5.1f}%")
        print(f"  Composite (match+partial / expected):             {(grand_match + grand_partial) / grand_exp * 100:5.1f}%")
    print()

    # Per-race table — only races that exist in the original
    print(f"{'RACE':45} {'HIT':>4} {'PART':>5} {'MISS':>5} {'BLANK':>6}")
    print("-" * 78)
    for race, st in sorted(per_race_summary.items()):
        if not (st["hit"] + st["miss"] + st["partial"] + st["blank"]):
            continue
        print(f"{race[:45]:45} {st['hit']:>4} {st['partial']:>5} {st['miss']:>5} {st['blank']:>6}")
    print()


if __name__ == "__main__":
    if len(sys.argv) < 3:
        # Default paths for this project
        root = Path(__file__).resolve().parent.parent
        ours = sorted((root / "data" / "reports").glob("*March*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)[0]
        orig = root / "Monthly Global Races Report_March 2026_14.04.2026.xlsx"
    else:
        ours = Path(sys.argv[1])
        orig = Path(sys.argv[2])
    score(ours, orig)
